import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment

def create_xlsreport(self):
    file_path = os.path.expanduser(self.entry.get().strip())
    report_data = []
    error_data = []
    for root, _, files in os.walk(file_path):
        for file in files:
            if file.endswith('.txt'):
                filepath = os.path.join(root, file)
                unit = file.replace("_report.txt", "")
                parsed_data, errors = parse_textfile(filepath)
                report_entry = {k: v for k, v in parsed_data.items() if not k.startswith("Desig ")}
                report_entry["Unit"] = unit
                report_data.append(report_entry)
                error_data.extend(errors)
    df = pd.DataFrame(report_data).drop_duplicates()
    error_df = pd.DataFrame(error_data, columns=["Unit", "Error Reference"])
    columns_order = ['Unit'] + [col for col in df.columns if col != 'Unit']
    df = df[columns_order]
    df = df.loc[:, ~df.columns.str.startswith("Check desig:")]
    output_excel = os.path.join(file_path, 'Lexology_XMLCleanup_Report.xlsx')
    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Report")
        error_df.to_excel(writer, index=False, sheet_name="Reference")
        report_sheet = writer.sheets["Report"]
        format_sheet(report_sheet)
        if not error_df.empty:
            error_sheet = writer.sheets["Reference"]
            format_sheet1(error_sheet)
    print("Consolidated report in Excel saved to", output_excel)

def parse_textfile(filepath):
    with open(filepath, 'r', encoding='utf-8') as file:
        content = file.read()
    patterns = {
        "Xml clean up:": r"Xml clean up:\s+(\d+)",
        "Contributing Editor(s):": r"Contributing Editor\(s\):\s+(\d+)",
        "By-line change:": r"By-line change:\s+(\d+)",
        "Contributing Editor(s) still in core:list:": r"Contributing Editor\(s\) still in core:list:\s+(\d+)",
        "Unique year in volnum:": r"Unique year in volnum:\s+(\d+)",
        "Volnum Year:": r"Volnum Year:\s+(\d+)",
        "Title Year:": r"Title Year:\s+(\d+)",
        "Non matching Country:": r"Non matching Country:\s*\[\('(.*?)'\)\]",
        "Non matching Question:": r"Non matching Question:\s*\[\('(.*?)'\)\]",
        "Number of accented character(s):": r"Number of accented character(s):\s+(\d+)",
        "Total tables found:": r"Total tables found:\s+(\d+)",
        "Modified tables:": r"Modified tables:\s+(\d+)",
        "Check desig:": r"Check desig:\s+\[(.*?)\]",
        "Core desig mismatch:": r"Core desig mismatch:\s+(\d+)",
    }
    data = {}
    errors = []
    unit_name = os.path.basename(filepath).replace("_report.txt", "")
    for category, pattern in patterns.items():
        matches = re.findall(pattern, content)

        if category == "Check desig:" and matches:
            desig_entries = eval(matches[0])  
            for entry in desig_entries:
                value = entry.get('value', '')
                text = entry.get('text', '')
                status = entry.get('status', '')
                data[f"Desig {value}"] = status
                errors.append({
                    "Unit": unit_name,
                    "Error Reference": f"Desig Value: {value}, Desig Text: {text}, Status: {status}"
                })
                print(f"Desig Value: {value}, Desig Text: {text}, Status: {status}")
        else:
            try:
                data[category] = int(matches[0]) if matches else 0
            except ValueError:
                print(f"Warning: Unable to convert {matches[0]} to an integer for category {category}.")
                data[category] = 0
    return data, errors

def format_sheet(sheet):
    header_font = Font(bold=True, color="FFFFFF")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = openpyxl.styles.PatternFill(start_color="A9A9A9", end_color="000000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    sheet.column_dimensions["A"].width = 50
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 12
    sheet.column_dimensions["E"].width = 12
    sheet.column_dimensions["F"].width = 12
    sheet.column_dimensions["G"].width = 12
    sheet.column_dimensions["H"].width = 10
    sheet.column_dimensions["I"].width = 12
    sheet.column_dimensions["J"].width = 12
    sheet.column_dimensions["K"].width = 12
    sheet.column_dimensions["L"].width = 12
    sheet.column_dimensions["M"].width = 12
    sheet.column_dimensions["N"].width = 13
    for row in sheet.iter_rows(min_row=2): 
        for cell in row:
            if cell.column_letter == "A":
                cell.alignment = Alignment(horizontal="left", wrap_text=True) 
            else:
                cell.alignment = Alignment(horizontal="center", wrap_text=True) 

def format_sheet1(sheet1):
    header_font = Font(bold=True, color="FFFFFF")
    for cell in sheet1[1]:
        cell.font = header_font
        cell.fill = openpyxl.styles.PatternFill(start_color="A9A9A9", end_color="000000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet1.column_dimensions["A"].width = 50
    sheet1.column_dimensions["B"].width = 80
    for col in sheet1.columns:
        for cell in col:
            cell.alignment = Alignment(wrap_text=True)